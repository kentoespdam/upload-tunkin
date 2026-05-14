"""Unit tests for KPISheetParser — no FastAPI/DB dependency.

Generates Excel fixtures in-memory via openpyxl.
Layout: row 1-4 = arbitrary filler, row 5 = header, row 6+ = data.

Run: uv run python test_kpi_sheet_parser.py
"""
import io

from app.tunkin.services import KPISheetParser
from app.tunkin.schemas import KPIRecord


# Columns wajib sesuai TEMPLATE_COLUMNS
_FULL_HEADERS = ["NO", "PERIODE", "NIPAM", "NAMA", "JUMLAH PENERIMAAN", "PPH21 TER"]


def _make_excel_bytes(
    rows: list[dict],
    headers: list[str] | None = None,
    case_insensitive: bool = False,
) -> bytes:
    """Create in-memory Excel file with rows 1-4 dummy, header at row 5, data row 6+."""
    import openpyxl
    from openpyxl.utils import get_column_letter

    if headers is None:
        headers = list(rows[0].keys()) if rows else _FULL_HEADERS

    if case_insensitive:
        headers = [h.lower() for h in headers]

    wb = openpyxl.Workbook()
    ws = wb.active

    # Rows 1-4: dummy filler (title, blank, etc.)
    for r in range(1, 5):
        ws[f"A{r}"] = f"dummy row {r}"

    # Row 5 (index 4 0-based): header
    for col_idx, h in enumerate(headers, 1):
        ws[f"{get_column_letter(col_idx)}5"] = h

    # Row 6+: data
    for row_idx, row in enumerate(rows, 6):
        for col_idx, h in enumerate(headers, 1):
            val = row.get(h) if not case_insensitive else row.get(h, row.get(h.upper()))
            ws[f"{get_column_letter(col_idx)}{row_idx}"] = val

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def test_valid_file_returns_records():
    parser = KPISheetParser()
    data = _make_excel_bytes([
        {"NO": 1, "PERIODE": 2501, "NIPAM": 12345678, "NAMA": "Alice", "JUMLAH PENERIMAAN": 500000, "PPH21 TER": 25000},
        {"NO": 2, "PERIODE": 2501, "NIPAM": 87654321, "NAMA": "Bob", "JUMLAH PENERIMAAN": 750000, "PPH21 TER": 37500},
    ])
    records = parser.parse(data)
    assert len(records) == 2
    assert all(isinstance(r, KPIRecord) for r in records)
    assert records[0].periode == "002501"
    assert records[0].nipam == "012345678"
    assert records[0].nama == "Alice"
    assert records[0].tunkin == 500000
    assert records[0].pph21_ter == 25000
    assert records[1].periode == "002501"
    assert records[1].nipam == "087654321"
    assert records[1].nama == "Bob"
    assert records[1].tunkin == 750000


def test_empty_file_raises():
    parser = KPISheetParser()
    try:
        parser.parse(b"")
        assert False, "should have raised"
    except Exception as e:
        from fastapi import HTTPException
        assert isinstance(e, HTTPException)
        assert "Terjadi kesalahan" in e.detail


def test_missing_column_raises():
    parser = KPISheetParser()
    data = _make_excel_bytes(
        [{"NO": 1, "PERIODE": 2501, "NAMA": "X"}],
        headers=["NO", "PERIODE", "NAMA"],  # sengaja hilang NIPAM, JUMLAH PENERIMAAN, PPH21 TER
    )
    try:
        parser.parse(data)
        assert False
    except Exception as e:
        from fastapi import HTTPException
        assert isinstance(e, HTTPException)
        assert "Kolom" in e.detail
        # required_lower lowercase → cari "nipam" bukan "NIPAM"
        assert "nipam" in e.detail


def test_zfill_applied():
    parser = KPISheetParser()
    data = _make_excel_bytes([
        {"NO": 1, "PERIODE": 1, "NIPAM": 12, "NAMA": "X", "JUMLAH PENERIMAAN": 100, "PPH21 TER": 5},
    ])
    records = parser.parse(data)
    assert records[0].periode == "000001"
    assert records[0].nipam == "000000012"


def test_case_insensitive_columns():
    """Header di Excel pakai lowercase → tetap dikenali."""
    parser = KPISheetParser()
    data = _make_excel_bytes([
        {"NO": 1, "PERIODE": 2501, "NIPAM": 12345678, "NAMA": "Alice",
         "JUMLAH PENERIMAAN": 500000, "PPH21 TER": 25000},
        {"NO": 2, "PERIODE": 2501, "NIPAM": 87654321, "NAMA": "Bob",
         "JUMLAH PENERIMAAN": 750000, "PPH21 TER": 37500},
    ], case_insensitive=True)
    records = parser.parse(data)
    assert len(records) == 2
    assert records[0].nama == "Alice"
    assert records[1].tunkin == 750000


def test_empty_data_after_header_raises():
    """Header ada tapi tanpa baris data → error."""
    parser = KPISheetParser()
    data = _make_excel_bytes([])  # hanya header, tanpa data
    try:
        parser.parse(data)
        assert False
    except Exception as e:
        from fastapi import HTTPException
        assert isinstance(e, HTTPException)
        assert "File Excel kosong" in e.detail


if __name__ == "__main__":
    tests = [
        test_valid_file_returns_records,
        test_empty_file_raises,
        test_missing_column_raises,
        test_zfill_applied,
        test_case_insensitive_columns,
        test_empty_data_after_header_raises,
    ]
    for t in tests:
        t()
        print(f"PASS: {t.__name__}")
    print("\nAll KPISheetParser unit tests passed!")
