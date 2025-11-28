import io
from typing import Optional, Annotated

from fastapi import UploadFile, HTTPException, Depends
from openpyxl.reader.excel import load_workbook

from app.core.config import Config, LOGGER
from app.core.databases import DatabaseHelper
from app.models.request_model import TunkinRequest, TunkinUploadRequest
from app.models.response_model import User
from app.repositories.sys_user import TokenHelper

TEMPLATE_COLUMN = [
    "NO",
    "PERIODE",
    "NIPAM",
    "JUMLAH PENERIMAAN"
]
START_ROW = 6

token_helper = TokenHelper()


def get_current_active_user(
        current_user: Annotated[User, Depends(token_helper.get_current_user)],
) -> User:
    if current_user['disabled']:
        raise HTTPException(status_code=400, detail="Inactive user")
    return current_user


class TunkinRepository:
    def __init__(self, config=Config(), db_helper: DatabaseHelper = DatabaseHelper()):
        self.config = config
        self.periode: str = ""
        self.file: Optional[UploadFile] = None
        self._allowed_extension = {'xlsx', 'xls'}
        self._max_file_size = 50 * 1024 * 1024  # 50MB
        self.db_helper = db_helper

    def fetch_page_data(self, periode: str, req: TunkinRequest, ):
        query = f"""
            SELECT
                kpi.id AS id,
                kpi.periode AS periode, 
                kpi.nipam AS nipam, 
                ep.emp_name AS nama, 
                po.pos_name AS jabatan, 
                org.org_name AS organisasi, 
                sef.text AS status_pegawai,
                kpi.nominal AS nominal
            FROM salary_kpi kpi
            INNER JOIN employee AS em ON kpi.nipam = em.emp_code
            INNER JOIN emp_profile AS ep ON em.emp_profile_id = ep.emp_profile_id
            INNER JOIN position AS po ON em.emp_pos_id = po.pos_id
            INNER JOIN organization AS org ON po.pos_org_id = org.org_id
            INNER JOIN sys_reference AS sef ON sef.`code` = 'emp_flag' 
                AND em.emp_flag = sef.`value` 
            WHERE kpi.periode = %s
        """
        params = (periode,)
        if req.nipam:
            query += " AND kpi.nipam = %s"
            params += (req.nipam,)
        if req.nama:
            query += " AND ep.emp_name LIKE %s"
            params += (f"%{req.nama}%",)

        query += " ORDER BY org.org_level , po.pos_level"
        LOGGER.info(query % params)
        return self.db_helper.fetch_page(query, params, req.page, req.size)

    async def check_exist_tunkin(self, periode: str):
        query = "SELECT EXISTS(SELECT 1 FROM salary_kpi WHERE periode = %s) AS is_exist"
        params = (periode,)
        try:
            result = self.db_helper.fetchone(query, params)
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

        if not result:
            return {"is_exist": False}

        return {"is_exist": bool(result.get("is_exist"))}

    async def upload(self, req: TunkinUploadRequest):
        self.cleanup()
        self.periode = req.periode
        self.file = req.file
        await self._file_checker()

        return await self.process_excel_data()

    async def _file_checker(self):
        """Validasi file sebelum diproses"""
        if not self.file:
            raise HTTPException(status_code=400, detail="File tidak ditemukan")

        if not self.file.filename:
            raise HTTPException(status_code=400, detail="Nama File tidak valid")

        await self._validate_file_extension()
        await self._validate_file_size()
        await self._validate_content_type()

    async def _validate_file_extension(self):
        file_extension = self.file.filename.lower().split('.')[-1]
        if file_extension not in self._allowed_extension:
            raise HTTPException(
                status_code=400,
                detail=f"Ekstensi file tidak diizinkan. Hanya ekstensi {', '.join(self._allowed_extension)} yang diperbolehkan."
            )

    async def _validate_file_size(self):
        if self.file.size > self._max_file_size:
            raise HTTPException(
                status_code=400,
                detail=f"Ukuran file melebihi batas maksimum {self._max_file_size / (1024 * 1024)} MB."
            )

        if self.file.size == 0:
            raise HTTPException(
                status_code=400,
                detail="File Kosong"
            )

    async def _validate_content_type(self):
        allowed_content_types = {
            'application/vnd.ms-excel',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
        if self.file.content_type not in allowed_content_types:
            raise HTTPException(
                status_code=400,
                detail="Tipe konten file tidak valid untuk file Excel."
            )

    async def process_excel_data(self):
        if not self.file:
            raise HTTPException(status_code=400, detail="File tidak ditemukan")

        try:
            await self.file.seek(0)

            contents = await self.file.read()
            file_like = io.BytesIO(contents)

            workbook = load_workbook(file_like)
            sheet = workbook.active

            data = []
            for row in sheet.iter_rows(min_row=START_ROW, values_only=True):
                if row is None or row[1] is None:
                    continue

                data.append((str(row[1]), str(row[2]), row[4]))

            if len(data) == 0:
                raise HTTPException(status_code=400, detail="File Excel kosong")

            unique_periode = set([row[0] for row in data])
            if len(unique_periode) > 1:
                raise HTTPException(status_code=400, detail="Periode pada file excel tidak konsisten!")

            for value in unique_periode:
                if value != self.periode:
                    raise HTTPException(status_code=400, detail="Periode pada file dengan request tidak sesuai!")

            query = f"""
                  INSERT INTO salary_kpi (periode, nipam, nominal)
                  VALUES (%s, %s, %s)
                  ON DUPLICATE KEY
                  UPDATE nominal = VALUES (nominal);
                  """

            affected = self.db_helper.save_update(query, data)
            processed_data = {
                "status": "success",
                "affected_rows": affected or 0
            }

            return processed_data
        except Exception as e:
            raise
        finally:
            if hasattr(self.file, 'seek'):
                await self.file.seek(0)

    def cleanup(self):
        if self.file:
            self.file.file.close()
            self.file = None


"""
    async def process_excel_data(self):
        if not self.file:
            raise HTTPException(status_code=400, detail="File tidak ditemukan")

        try:
            await self.file.seek(0)

            contents = await self.file.read()
            file_like = io.BytesIO(contents)

            df = pd.read_excel(file_like)

            if df.empty:
                raise HTTPException(status_code=400, detail="File Excel kosong")

            for column in TEMPLATE_COLUMN:
                if column not in df.columns:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Kolom '{column}' tidak ditemukan dalam file Excel."
                    )

            df["PERIODE"] = df["PERIODE"].astype(str).str.zfill(6)
            df["NIPAM"] = df["NIPAM"].astype(str).str.zfill(8)
            data = [(
                row['PERIODE'],
                row['NIPAM'],
                row['JUMLAH PENERIMAAN']
            ) for _, row in df.iterrows()]

            query = f"
                  INSERT INTO salary_kpi (periode, nipam, nominal)
                  VALUES (%s, %s, %s)
                  ON DUPLICATE KEY
                  UPDATE
                      nominal =
                  VALUES (nominal);
                  "

            affected = self.db_helper.save_update(query, data)
            processed_data = {
                "status": "success",
                "affected_rows": affected or 0
            }

            return processed_data
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Terjadi kesalahan saat memproses file Excel: {str(e)}")
        finally:
            if hasattr(self.file, 'seek'):
                await self.file.seek(0)
"""
