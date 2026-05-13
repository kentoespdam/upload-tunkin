"""Unit tests for KPISheetParser — no FastAPI/DB dependency.

Generates Excel fixtures in-memory via openpyxl.

Run: uv run python test_kpi_sheet_parser.py
"""
import io

from app.services.kpi_sheet_parser import KPISheetParser
from app.models.kpi import KPIRecord


def _make_excel_bytes(rows: list[dict]) -> bytes:
    """Create an in-memory Excel file from a list of row dicts."""
    import openpyxl
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active

    # Header row
    headers = list(rows[0].keys()) if rows else ["NO", "PERIODE", "NIPAM", "JUMLAH PENERIMAAN"]
    for col_idx, h in enumerate(headers, 1):
        ws[f"{get_column_letter(col_idx)}1"] = h

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, h in enumerate(headers, 1):
            ws[f"{get_column_letter(col_idx)}{row_idx}"] = row[h]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def test_valid_file_returns_records():
    parser = KPISheetParser()
    data = _make_excel_bytes([
        {"NO": 1, "PERIODE": 2501, "NIPAM": 12345678, "JUMLAH PENERIMAAN": 500000},
        {"NO": 2, "PERIODE": 2501, "NIPAM": 87654321, "JUMLAH PENERIMAAN": 750000},
    ])
    records = parser.parse(data)
    assert len(records) == 2
    assert all(isinstance(r, KPIRecord) for r in records)
    assert records[0].periode == "002501"  # zfilled to 6
    assert records[0].nipam == "12345678"
    assert records[0].nominal == 500000
    assert records[1].periode == "002501"


def test_empty_file_raises():
    parser = KPISheetParser()
    try:
        parser.parse(b"")
        assert False, "should have raised"
    except Exception as e:
        from fastapi import HTTPException
        assert isinstance(e, HTTPException)
        # Empty bytes fails pandas with "file format cannot be determined"
        # which maps to 500 "Terjadi kesalahan..."
        assert "Terjadi kesalahan" in e.detail


def test_missing_column_raises():
    parser = KPISheetParser()
    data = _make_excel_bytes([
        {"NO": 1, "PERIODE": 2501},  # missing NIPAM, JUMLAH PENERIMAAN
    ])
    try:
        parser.parse(data)
        assert False
    except Exception as e:
        from fastapi import HTTPException
        assert isinstance(e, HTTPException)
        assert "Kolom" in e.detail
        assert "NIPAM" in e.detail


def test_zfill_applied():
    parser = KPISheetParser()
    data = _make_excel_bytes([
        {"NO": 1, "PERIODE": 1, "NIPAM": 12, "JUMLAH PENERIMAAN": 100},
    ])
    records = parser.parse(data)
    assert records[0].periode == "000001"  # 6 chars
    assert records[0].nipam == "00000012"   # 8 chars


if __name__ == "__main__":
    tests = [test_valid_file_returns_records, test_empty_file_raises,
             test_missing_column_raises, test_zfill_applied]
    for t in tests:
        t()
        print(f"PASS: {t.__name__}")
    print("\nAll KPISheetParser unit tests passed!")
